from openpyxl import *
from tkinter import *

wb = load_workbook('Bai6.xlsx')
sheet = wb.active
def excel():
	
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 50
	sheet.column_dimensions['C'].width = 30
	sheet.column_dimensions['D'].width = 50
	sheet.column_dimensions['E'].width = 15
	sheet.column_dimensions['F'].width = 15
	sheet.column_dimensions['G'].width = 30
	sheet.column_dimensions['H'].width = 30

	sheet.cell(row=1, column=1).value = "Mã số sinh viên"
	sheet.cell(row=1, column=2).value = "Họ tên"
	sheet.cell(row=1, column=3).value = "Ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Số điện thoại"
	sheet.cell(row=1, column=6).value = "Học kỳ"
	sheet.cell(row=1, column=7).value = "Năm học"
	sheet.cell(row=1, column=8).value = "Môn học"

def focus0(event):
	mssv_field.focus_set()
def focus1(event):
	ho_ten_field.focus_set()
def focus2(event):
	ngay_Sinh_field.focus_set()
def focus3(event):
	ngay_Sinh_field.focus_set()
def focus4(event):
	so_DT_field.focus_set()
def focus5(event):
	hoc_ky_field.focus_set()
def focus6(event):
	nam_hoc_field.focus_set()
def focus7(event):
	mon_hoc_field.focus_set()

def clear():
	mssv_field.delete(0, END)
	ho_ten_field.delete(0, END)
	ngay_Sinh_field.delete(0, END)
	ngay_Sinh_field.delete(0, END)
	so_DT_field.delete(0, END)
	hoc_ky_field.delete(0, END)
	nam_hoc_field.delete(0, END)
	mon_hoc_field.delete(0, END)

def insert():
	
	if (mssv_field.get() == "" and
		ho_ten_field.get() == "" and
		ngay_Sinh_field.get() == "" and
		email_field.get() == "" and
		so_DT_field.get() == "" and
		hoc_ky_field.get() == "" and
		nam_hoc_field.get() == "" and
  		mon_hoc_field.get() == ""):
		print("empty input")
	else:
		current_row = sheet.max_row
		current_column = sheet.max_column

		sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
		sheet.cell(row=current_row + 1, column=2).value = ho_ten_field.get()
		sheet.cell(row=current_row + 1, column=3).value = ngay_Sinh_field.get()
		sheet.cell(row=current_row + 1, column=4).value = email_field.get()
		sheet.cell(row=current_row + 1, column=5).value = so_DT_field.get()
		sheet.cell(row=current_row + 1, column=6).value = hoc_ky_field.get()
		sheet.cell(row=current_row + 1, column=7).value = nam_hoc_field.get()
		sheet.cell(row=current_row + 1, column=7).value = mon_hoc_field.get()
		wb.save('Bai6.xlsx')
		mssv_field.focus_set()
	clear()


# Driver code
if __name__ == "__main__":
	
	root = Tk()
	root.configure(background='light green')
	root.title("Đăng kí học phần")
	root.geometry("500x300")

	heading = Label(root, text="THÔNG TIN ĐĂNG KÍ HỌC PHẦN", bg="aqua")
	mssv = Label(root, text="Mã số sinh viên", bg="light green")
	ho_ten = Label(root, text="Họ tên", bg="light green")
	ngay_Sinh = Label(root, text="Ngày sinh", bg="light green")
	email = Label(root, text="Email", bg="light green")
	so_DT = Label(root, text="Số điện thoại", bg="light green")
	hoc_ky = Label(root, text="Học kỳ", bg="light green")
	nam_hoc = Label(root, text="Năm học", bg="light green")

	exit_button = Button(root, text="Thoát", fg="Black", bg="Yellow", command=root.quit).place(x=250 ,y=250)

	heading.grid(row=0, column=1)
	mssv.grid(row=1, column=0,sticky="w")
	ho_ten.grid(row=2, column=0,sticky="w")
	ngay_Sinh.grid(row=3, column=0,sticky="w")
	email.grid(row=4, column=0,sticky="w")
	so_DT.grid(row=5, column=0,sticky="w")
	hoc_ky.grid(row=6, column=0,sticky="w")
	nam_hoc.grid(row=7, column=0,sticky="w")

	mssv_field = Entry(root)
	ho_ten_field = Entry(root)
	ngay_Sinh_field = Entry(root)
	email_field = Entry(root)
	so_DT_field = Entry(root)
	hoc_ky_field = Entry(root)
	nam_hoc_field = Entry(root)	
	mon_hoc_field = Entry(root)

	mssv_field.bind("<Return>", focus1)
	ho_ten_field.bind("<Return>", focus2)
	ngay_Sinh_field.bind("<Return>", focus3)
	email_field.bind("<Return>", focus4)
	so_DT_field.bind("<Return>", focus5)
	nam_hoc_field.bind("<Return>", focus6)
	mon_hoc_field.bind("<Return>", focus7)
 

	mssv_field.grid(row=1, column=1, ipadx="100")
	ho_ten_field.grid(row=2, column=1, ipadx="100")
	ngay_Sinh_field.grid(row=3, column=1, ipadx="100")
	email_field.grid(row=4, column=1, ipadx="100")
	so_DT_field.grid(row=5, column=1, ipadx="100")
	nam_hoc_field.grid(row=6, column=1, ipadx="100")
	mon_hoc_field.grid(row=7, column=1, ipadx="100")

	excel()
	programming_label = Label(root, text="Chọn môn học", bg="light green").place(x= 0,y=170)


	selected_language1 = StringVar()
	selected_language2 = StringVar()
	selected_language3 = StringVar()
	selected_language4 = StringVar()

	language1 = Checkbutton(root, text="Lập trình Python", variable=selected_language1, bg="light green").place(x= 83,y=170)
	language2 = Checkbutton(root, text="Lập trình Java", variable=selected_language2, bg="light green").place(x= 240,y=170)
	language3 = Checkbutton(root, text="Công nghệ phần mềm", variable=selected_language3, bg="light green").place(x= 83,y=210)
	language4 = Checkbutton(root, text="Phát triển ứng dựng web", variable=selected_language4, bg="light green").place(x= 240,y=210)

	submit = Button(root, text="Đăng kí", fg="Black",bg="Yellow", command=insert).place(x=150 ,y=250)

	root.mainloop()
